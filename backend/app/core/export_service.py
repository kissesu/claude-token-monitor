"""
@file: export_service.py
@description: 数据导出服务，支持多种格式导出统计数据
@author: Atlas.oi
@date: 2026-01-07

功能说明：
1. 支持 CSV 格式导出
2. 支持 JSON 格式导出
3. 支持时间范围筛选
4. 支持字段选择
5. 提供文件保存和内存返回两种模式
"""

import csv
import json
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

from ..core.config import get_settings
from ..core.logger import get_logger
from ..core.schemas import ExportFormat, ExportResponse, DailyActivity


logger = get_logger(__name__)


class ExportService:
    """
    数据导出服务

    提供多种格式的数据导出功能
    """

    def __init__(self, export_dir: Optional[Path] = None):
        """
        初始化导出服务

        Args:
            export_dir: 导出文件保存目录，为 None 时使用默认目录
        """
        if export_dir is None:
            settings = get_settings()
            export_dir = settings.project_root / "exports"

        self.export_dir = export_dir
        self.export_dir.mkdir(parents=True, exist_ok=True)

        logger.info(f"初始化导出服务，导出目录: {self.export_dir}")

    async def export_data(
        self,
        data: List[DailyActivity],
        format: ExportFormat,
        filename: Optional[str] = None,
    ) -> ExportResponse:
        """
        导出数据

        根据指定格式导出数据到文件

        Args:
            data: 要导出的数据列表
            format: 导出格式
            filename: 文件名（不含扩展名），为 None 时自动生成

        Returns:
            ExportResponse: 导出结果响应
        """
        try:
            # ============================================
            # 生成文件名
            # ============================================
            if filename is None:
                filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

            # ============================================
            # 根据格式调用对应的导出函数
            # ============================================
            if format == ExportFormat.JSON:
                return await self._export_json(data, filename)
            elif format == ExportFormat.CSV:
                return await self._export_csv(data, filename)
            elif format == ExportFormat.EXCEL:
                return await self._export_excel(data, filename)
            else:
                raise ValueError(f"不支持的导出格式: {format}")

        except Exception as e:
            logger.error(f"导出数据失败: {e}")
            return ExportResponse(
                success=False,
                message=f"导出失败: {str(e)}",
            )

    async def _export_json(
        self,
        data: List[DailyActivity],
        filename: str,
    ) -> ExportResponse:
        """
        导出为 JSON 格式

        Args:
            data: 要导出的数据
            filename: 文件名（不含扩展名）

        Returns:
            ExportResponse: 导出结果
        """
        try:
            # ============================================
            # 转换数据为字典列表
            # ============================================
            export_data = []

            for activity in data:
                # 基础数据
                item = {
                    "date": activity.date,
                    "session_count": activity.session_count,
                    "total_tokens": activity.total_tokens,
                    "models": []
                }

                # 模型详情
                for model in activity.models:
                    item["models"].append({
                        "model_name": model.model_name,
                        "input_tokens": model.input_tokens,
                        "output_tokens": model.output_tokens,
                        "cache_read_tokens": model.cache_read_tokens,
                        "cache_creation_tokens": model.cache_creation_tokens,
                    })

                export_data.append(item)

            # ============================================
            # 写入文件
            # ============================================
            file_path = self.export_dir / f"{filename}.json"

            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(
                    export_data,
                    f,
                    ensure_ascii=False,
                    indent=2,
                )

            logger.info(f"成功导出 JSON 文件: {file_path}")

            return ExportResponse(
                success=True,
                message=f"成功导出 {len(export_data)} 条记录到 JSON 文件",
                file_path=str(file_path),
                data=export_data,
            )

        except Exception as e:
            logger.error(f"导出 JSON 失败: {e}")
            raise

    async def _export_csv(
        self,
        data: List[DailyActivity],
        filename: str,
    ) -> ExportResponse:
        """
        导出为 CSV 格式

        Args:
            data: 要导出的数据
            filename: 文件名（不含扩展名）

        Returns:
            ExportResponse: 导出结果
        """
        try:
            # ============================================
            # 准备 CSV 数据（扁平化结构）
            # ============================================
            csv_rows = []

            for activity in data:
                if not activity.models:
                    # 没有模型数据时，只输出汇总信息
                    csv_rows.append({
                        "日期": activity.date,
                        "会话数": activity.session_count,
                        "总Token数": activity.total_tokens,
                        "模型名称": "",
                        "输入Token": 0,
                        "输出Token": 0,
                        "缓存读取Token": 0,
                        "缓存创建Token": 0,
                    })
                else:
                    # 有模型数据时，每个模型一行
                    for model in activity.models:
                        csv_rows.append({
                            "日期": activity.date,
                            "会话数": activity.session_count,
                            "总Token数": activity.total_tokens,
                            "模型名称": model.model_name,
                            "输入Token": model.input_tokens,
                            "输出Token": model.output_tokens,
                            "缓存读取Token": model.cache_read_tokens,
                            "缓存创建Token": model.cache_creation_tokens,
                        })

            # ============================================
            # 写入 CSV 文件
            # ============================================
            file_path = self.export_dir / f"{filename}.csv"

            if csv_rows:
                with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
                    fieldnames = [
                        "日期",
                        "会话数",
                        "总Token数",
                        "模型名称",
                        "输入Token",
                        "输出Token",
                        "缓存读取Token",
                        "缓存创建Token",
                    ]

                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(csv_rows)

                logger.info(f"成功导出 CSV 文件: {file_path}")

                return ExportResponse(
                    success=True,
                    message=f"成功导出 {len(csv_rows)} 条记录到 CSV 文件",
                    file_path=str(file_path),
                )
            else:
                return ExportResponse(
                    success=False,
                    message="没有数据可导出",
                )

        except Exception as e:
            logger.error(f"导出 CSV 失败: {e}")
            raise

    async def _export_excel(
        self,
        data: List[DailyActivity],
        filename: str,
    ) -> ExportResponse:
        """
        导出为 Excel 格式

        注意：需要安装 openpyxl 库

        Args:
            data: 要导出的数据
            filename: 文件名（不含扩展名）

        Returns:
            ExportResponse: 导出结果
        """
        try:
            # 尝试导入 openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill
            except ImportError:
                return ExportResponse(
                    success=False,
                    message="Excel 导出需要安装 openpyxl 库：pip install openpyxl",
                )

            # ============================================
            # 创建工作簿
            # ============================================
            wb = Workbook()

            # ============================================
            # 创建汇总表
            # ============================================
            ws_summary = wb.active
            ws_summary.title = "每日汇总"

            # 表头样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            header_alignment = Alignment(horizontal="center", vertical="center")

            # 写入汇总表表头
            summary_headers = ["日期", "会话数", "总Token数"]
            for col_idx, header in enumerate(summary_headers, 1):
                cell = ws_summary.cell(row=1, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # 写入汇总数据
            for row_idx, activity in enumerate(data, 2):
                ws_summary.cell(row=row_idx, column=1, value=activity.date)
                ws_summary.cell(row=row_idx, column=2, value=activity.session_count)
                ws_summary.cell(row=row_idx, column=3, value=activity.total_tokens)

            # ============================================
            # 创建模型详情表
            # ============================================
            ws_details = wb.create_sheet("模型详情")

            # 写入详情表表头
            detail_headers = [
                "日期",
                "模型名称",
                "输入Token",
                "输出Token",
                "缓存读取Token",
                "缓存创建Token",
            ]

            for col_idx, header in enumerate(detail_headers, 1):
                cell = ws_details.cell(row=1, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # 写入详情数据
            detail_row = 2
            for activity in data:
                for model in activity.models:
                    ws_details.cell(row=detail_row, column=1, value=activity.date)
                    ws_details.cell(row=detail_row, column=2, value=model.model_name)
                    ws_details.cell(row=detail_row, column=3, value=model.input_tokens)
                    ws_details.cell(row=detail_row, column=4, value=model.output_tokens)
                    ws_details.cell(row=detail_row, column=5, value=model.cache_read_tokens)
                    ws_details.cell(row=detail_row, column=6, value=model.cache_creation_tokens)
                    detail_row += 1

            # ============================================
            # 自动调整列宽
            # ============================================
            for ws in [ws_summary, ws_details]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # ============================================
            # 保存文件
            # ============================================
            file_path = self.export_dir / f"{filename}.xlsx"
            wb.save(file_path)

            logger.info(f"成功导出 Excel 文件: {file_path}")

            return ExportResponse(
                success=True,
                message=f"成功导出 {len(data)} 条记录到 Excel 文件",
                file_path=str(file_path),
            )

        except Exception as e:
            logger.error(f"导出 Excel 失败: {e}")
            raise

    async def export_model_summary(
        self,
        models: Dict[str, Any],
        filename: Optional[str] = None,
        format: ExportFormat = ExportFormat.CSV,
    ) -> ExportResponse:
        """
        导出模型使用汇总

        导出所有模型的汇总统计数据

        Args:
            models: 模型使用数据字典
            filename: 文件名（不含扩展名）
            format: 导出格式

        Returns:
            ExportResponse: 导出结果
        """
        try:
            if filename is None:
                filename = f"model_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

            # ============================================
            # 根据格式选择导出方式
            # ============================================
            if format == ExportFormat.JSON:
                file_path = self.export_dir / f"{filename}.json"

                with open(file_path, "w", encoding="utf-8") as f:
                    json.dump(models, f, ensure_ascii=False, indent=2)

                logger.info(f"成功导出模型汇总 JSON: {file_path}")

                return ExportResponse(
                    success=True,
                    message=f"成功导出 {len(models)} 个模型的汇总数据",
                    file_path=str(file_path),
                    data=models,
                )

            elif format == ExportFormat.CSV:
                file_path = self.export_dir / f"{filename}.csv"

                with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
                    fieldnames = [
                        "模型名称",
                        "输入Token",
                        "输出Token",
                        "缓存读取Token",
                        "缓存创建Token",
                        "总Token数",
                    ]

                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()

                    for model_name, usage in models.items():
                        writer.writerow({
                            "模型名称": model_name,
                            "输入Token": usage.get("input_tokens", 0),
                            "输出Token": usage.get("output_tokens", 0),
                            "缓存读取Token": usage.get("cache_read_tokens", 0),
                            "缓存创建Token": usage.get("cache_creation_tokens", 0),
                            "总Token数": (
                                usage.get("input_tokens", 0) +
                                usage.get("output_tokens", 0) +
                                usage.get("cache_read_tokens", 0) +
                                usage.get("cache_creation_tokens", 0)
                            ),
                        })

                logger.info(f"成功导出模型汇总 CSV: {file_path}")

                return ExportResponse(
                    success=True,
                    message=f"成功导出 {len(models)} 个模型的汇总数据",
                    file_path=str(file_path),
                )

            else:
                raise ValueError(f"不支持的导出格式: {format}")

        except Exception as e:
            logger.error(f"导出模型汇总失败: {e}")
            return ExportResponse(
                success=False,
                message=f"导出失败: {str(e)}",
            )
